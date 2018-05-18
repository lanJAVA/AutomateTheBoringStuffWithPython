#! python3
# multiplicationTable.py - 12.13.1 乘法表

import sys, openpyxl
from openpyxl.styles import Font

if len(sys.argv) < 2:
    print('Usage: multiplicationTable.py number')
    sys.exit()

N = int(sys.argv[1])
boldFont = Font(bold=True)
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, N + 1):
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=i+1, column=1).font = boldFont
    for j in range(1, N + 1):
        sheet.cell(row=1, column=j+1).value = j
        sheet.cell(row=1, column=j + 1).font = boldFont
        sheet.cell(row=i+1, column=j+1).value = i * j
wb.save('multiplicationTable.xlsx')
