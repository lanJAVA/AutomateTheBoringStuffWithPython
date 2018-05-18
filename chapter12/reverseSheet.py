#! python
# reverseSheet.py - 12.13.3 电子表格单元格翻转程序

import openpyxl, sys

if len(sys.argv) != 2:
    print('Usage: reverseSheet.py fileName')
    sys.exit()

fileName = sys.argv[1]

wb = openpyxl.load_workbook(fileName)
sheet = wb.active
newBook = openpyxl.Workbook()
newSheet = newBook.active

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        newSheet.cell(row=j, column=i).value = sheet.cell(row=i, column=j).value

newBook.save('reverse' + fileName)