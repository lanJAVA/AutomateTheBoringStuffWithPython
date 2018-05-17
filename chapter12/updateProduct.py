#! python3
# updateProduct.py - Corrects costs in product sales spreadsheet.

import openpyxl


wb = openpyxl.load_workbook('productSales.xlsx')
sheet = wb['Sheet']
# The product types and their updated prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}
# TODO: Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row + 1):     # skip the first row
    productName = sheet.cell(row=rowNum, column=1).value
    if productName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[productName]

wb.save('updatedProductSales.xlsx')
