import openpyxl

# wb = openpyxl.load_workbook('example.xlsx')
# print(wb.sheetnames)        # wb.get_sheet_names()
# sheet = wb['Sheet3']        # wb.get_sheet_by_name('Sheet3')
# print(sheet)
# activeSheet = wb.active     # wb.get_active_sheet()
# cell = activeSheet['B1']
# print('Row', str(cell.row), ', Column', str(cell.column), ' is ' cell.value)
# print(type(cell.value))

# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.active
# cell = sheet.cell(row=1, column=2)
# print(cell.value)
# for i in range(1, 8, 2):
#     print(i, sheet.cell(row=i, column=2).value)
# print(sheet.max_row)
# print(sheet.max_column)
# print(cell.col_idx)

# from openpyxl import utils
# print(utils.column_index_from_string('Z'))
# print(utils.get_column_letter(45))
# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']
# for rowOfCellObjects in sheet['A1':'C3']:  # sheet['A1':'C3']得到一个Generator对象
#     for cellObj in rowOfCellObjects:
#         print(cellObj.coordinate, cellObj.value)
#     print('--- END OF ROW ---')
# print(tuple(sheet.columns)[1])
# for cellObj in list(sheet.columns)[1]:
#     print(cellObj.value)

# import census2010
# print(census2010.allData['AK']['Anchorage'])
# anchoragePop = census2010.allData['AK']['Anchorage']['pop']
# print('The 2010 population of Anchorage was ' + str(anchoragePop))

# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# sheet = wb.active
# print(sheet.title)
# sheet.title = 'Spam Bacon Eggs Sheet'
# print(wb.sheetnames)
# wb.save('example_copy.xlsx')

# wb.create_sheet()
# print(wb.sheetnames)
# wb.create_sheet(index=0, title='First Sheet')
# print(wb.sheetnames)
# wb.create_sheet(index=2, title='Middle Sheet')
# print(wb.sheetnames)
# del wb['Middle Sheet']
# print(wb.sheetnames)
# wb.remove(wb['Sheet1'])
# print(wb.sheetnames)
# sheet = wb.active
# sheet['A1'] = 'Hello World!'
# print(sheet['A1'].value)
# wb.save('copy.xlsx')

# from openpyxl.styles import Font
# wb = openpyxl.Workbook()
# sheet = wb['Sheet']
# italic24Font = Font(size=24, italic=True)
# sheet['A1'].font = italic24Font
# sheet['D3'].font = italic24Font
# sheet['A1'] = 'Hello world!'
# sheet['D3'] = '你好'
# wb.save('styled.xlsx')

# wb = openpyxl.Workbook()
# ws = wb.active
# ws['A1'] = 200
# ws['A2'] = 300
# ws['A3'] = '=SUM(A1:A2)'
# wb.save('writeFormula.xlsx')

# wb = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
# ws = wb.active
# print(ws['A3'].value)
# wb.close()
#
# wbo = openpyxl.load_workbook('writeFormula.xlsx')
# wso = wbo.active
# print(wso['A3'].value)
# wbo.close()

# wb = openpyxl.Workbook()
# ws = wb.active
# ws['A1'] = 'Tall row'
# ws['A2'] = 'Wide column'
# ws.row_dimensions[1].height = 70
# ws.column_dimensions['B'].width = 20
# wb.save('dimensions.xlsx')

# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.merge_cells('A1:D3')
# sheet['A1'] = 'Twelve cells merged together.'
# sheet.merge_cells('C5:D5')
# sheet['C5'] = 'Two merged cells.'
# wb.save('merged.xlsx')

# wb = openpyxl.load_workbook('merged.xlsx')
# sheet = wb.active
# sheet.unmerge_cells('A1:D3')
# sheet.unmerge_cells('C5:D5')
# wb.save('merged.xlsx')

# wb = openpyxl.load_workbook('productSales.xlsx')
# sheet = wb.active
# sheet.freeze_panes = 'B3'
# wb.save('freezeExample.xlsx')

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):
    sheet['A' + str(i)] = i
refObj = openpyxl.charts.Reference(sheet, (1, 1), (10, 1))
seriesObj = openpyxl.charts.Series(refObj, title='First series')
chartObj = openpyxl.charts.BarChart()
chartObj.append(seriesObj)
chartObj.drawing.top = 50
chartObj.drawing.left = 100
chartObj.drawing.width = 300
chartObj.drawing.height = 200
sheet.add_chart(chartObj)
wb.save('sampleChart.xlsx')